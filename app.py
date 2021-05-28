import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

data_wb = load_workbook("data.xlsx")
data_ws = data_wb.active
data_ws['T1'] = "CODE"
codeRow = data_ws['T']
for cell in codeRow:
    if cell.value != "CODE" or len(cell.value) != 14:
        cell.value = cell.value.split(" ", 1)[0]
data_wb.save('data.xlsx')

data_frame_raw = pd.read_excel("data.xlsx", index_col=0)
df_prices_raw = pd.read_excel("banggia.xlsx", index_col=0)
df_data = data_frame_raw.loc[data_frame_raw["Unnamed: 27"] == 0]
df_prices = df_prices_raw["ĐƠN GIÁ.1"]
data_FOC = pd.merge(df_data, df_prices, how='left', on='CODE')
data_FOC.to_excel("test.xlsx")

wb_data_FOC = load_workbook("test.xlsx")
ws_data_FOC = wb_data_FOC.active
sumCol = ws_data_FOC['AF']
for i, sumCell in enumerate(sumCol, 1):
    if i > 1 and ws_data_FOC[f"AE{i}"].value != None:
        sumCell.value = (ws_data_FOC[f"AE{i}"].value) * (ws_data_FOC[f"M{i}"].value)

idCustomerCol = ws_data_FOC['S']
for j, cellid in enumerate( idCustomerCol, 1 ):
    ws_data_FOC[f"Y{j}"] = cellid.value.split(" ", 1)[0]
    ws_data_FOC[f"Z{j}"] = cellid.value.split(" ", 1)[1]


ws_data_FOC['AF1'] = "Thành Tiền"
ws_data_FOC['Y1'] = "Code KH"
ws_data_FOC['Z1'] = "Tên KH"

wb_data_FOC.save("dataFOC.xlsx")

df_dataFOC = pd.read_excel("dataFOC.xlsx", index_col=0)

# Function filter companies from list
def checkName(dfSerie):
    arrBool = []
    for i in dfSerie:
        if "C«ng ty" in i:
            arrBool.append(False)
            continue
        elif "C«ng Ty" in i:
            arrBool.append(False)
            continue
        arrBool.append(True)
    return arrBool

df_dataFOC = df_dataFOC.loc[checkName(df_dataFOC["Tên KH"])]
pivot_FOC = pd.pivot_table(df_dataFOC, values="Thành Tiền", index="Code KH", aggfunc=np.sum)
pivot_FOC = pd.merge(pivot_FOC, pd.read_excel("dataFOC.xlsx", index_col=0, usecols="Y,Z"), how='left', on='Code KH')
pivot_FOC.to_excel("pivotFOC.xlsx")


os.remove("test.xlsx")